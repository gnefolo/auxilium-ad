"""
Estrae dalle tavole ufficiali ISTAT (Sistema di tavole Input-Output, anni 2020-2022,
63 branche NACE Rev.2 - https://www.istat.it/tavole-di-dati/il-sistema-di-tavole-input-output-anni-2020-2022/)
la matrice dei coefficienti tecnici DOMESTICI (tavola simmetrica totale meno la tavola
simmetrica delle importazioni) e il vettore dei coefficienti di valore aggiunto, e li
salva in un JSON compatto (`data/istat_io_2022.json`) usato da `multipliers.py`.

Va rieseguito solo se si vuole aggiornare l'anno di riferimento o rigenerare la cache
da zero. I file grezzi ISTAT sono in `data/raw/` (scaricati da istat.it il 21/08/2026).

Uso: python3 build_io_cache.py
"""

import json
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).parent / "data"
RAW_DIR = DATA_DIR / "raw"
YEAR = 2022  # ultimo anno disponibile nel rilascio 2020-2022

BRANCH_ROW_START = 7   # prima riga con un codice branca (V01)
BRANCH_ROW_END = 69    # ultima riga con un codice branca (VT) - 63 branche
BRANCH_COL_START = 3   # prima colonna con un codice branca (V01)
BRANCH_COL_END = 65    # ultima colonna con un codice branca (VT) - 63 branche
ROW_REDDITI_LAVORO_DIP = 74   # SIMM_TOT: "Redditi da lavoro dipendente" per branca (colonna)
ROW_RISULTATO_LORDO_GESTIONE = 79  # SIMM_TOT: "Risultato lordo di gestione" (proxy EBITDA settoriale)
ROW_VALORE_AGGIUNTO = 81
ROW_PRODUZIONE = 82

# USEPB_63B.xlsx (tavola degli impieghi ai prezzi base)
COL_SPESA_CONSUMI_FAMIGLIE = 67  # "Spesa per consumi finali delle famiglie" (per prodotto/riga)
ROW_ORE_LAVORATE = 89             # "Ore lavorate (migliaia)" per branca (colonna)

# Fonti (21/08/2026):
# - Cuneo fiscale e contributivo Italia, lavoratore medio single, 2022: 45,0%
#   (OECD, Taxing Wages 2023 - https://www.oecd.org/en/publications/taxing-wages-2023_2eec52a9.html)
# - Propensione al risparmio delle famiglie, 2022: 7,8%
#   (Istat, "I conti nazionali per settore istituzionale, anni 1995-2023", 5/4/2024,
#   https://www.istat.it/it/files/2024/04/Conti-istituzionali-1995-2023.pdf, pag. 4)
TAX_WEDGE_LAVORO_2022 = 0.450
PROPENSIONE_RISPARMIO_FAMIGLIE_2022 = 0.078
# Quota di un euro aggiuntivo di reddito da lavoro dipendente che si traduce in nuova
# spesa per consumi (il resto è cuneo fiscale/contributivo + risparmio):
PROPENSITA_MARGINALE_CONSUMO = (1 - TAX_WEDGE_LAVORO_2022) * (1 - PROPENSIONE_RISPARMIO_FAMIGLIE_2022)


def load_matrix(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    codes = []
    names = []
    for r in range(BRANCH_ROW_START, BRANCH_ROW_END + 1):
        codes.append(ws.cell(row=r, column=1).value.strip())
        names.append(ws.cell(row=r, column=2).value.strip())

    n = len(codes)
    matrix = [[0.0] * n for _ in range(n)]
    for i, r in enumerate(range(BRANCH_ROW_START, BRANCH_ROW_END + 1)):
        for j, c in enumerate(range(BRANCH_COL_START, BRANCH_COL_END + 1)):
            v = ws.cell(row=r, column=c).value
            matrix[i][j] = float(v) if v is not None else 0.0

    output = []
    value_added = []
    redditi_lavoro = []
    risultato_lordo_gestione = []
    for c in range(BRANCH_COL_START, BRANCH_COL_END + 1):
        o = ws.cell(row=ROW_PRODUZIONE, column=c).value
        va = ws.cell(row=ROW_VALORE_AGGIUNTO, column=c).value
        rl = ws.cell(row=ROW_REDDITI_LAVORO_DIP, column=c).value
        rlg = ws.cell(row=ROW_RISULTATO_LORDO_GESTIONE, column=c).value
        output.append(float(o) if o is not None else 0.0)
        value_added.append(float(va) if va is not None else 0.0)
        redditi_lavoro.append(float(rl) if rl is not None else 0.0)
        risultato_lordo_gestione.append(float(rlg) if rlg is not None else 0.0)

    wb.close()
    return codes, names, matrix, output, value_added, redditi_lavoro, risultato_lordo_gestione


def load_use_table_columns(path: Path, sheet: str):
    """Da USEPB_63B.xlsx: spesa per consumi finali delle famiglie (per prodotto) e ore
    lavorate in migliaia (per branca)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    household_consumption = []
    for r in range(BRANCH_ROW_START, BRANCH_ROW_END + 1):
        v = ws.cell(row=r, column=COL_SPESA_CONSUMI_FAMIGLIE).value
        household_consumption.append(float(v) if v is not None else 0.0)

    hours_worked_thousands = []
    for c in range(BRANCH_COL_START, BRANCH_COL_END + 1):
        v = ws.cell(row=ROW_ORE_LAVORATE, column=c).value
        hours_worked_thousands.append(float(v) if v is not None else 0.0)

    wb.close()
    return household_consumption, hours_worked_thousands


def main():
    tot_codes, tot_names, tot_matrix, output, value_added, redditi_lavoro, risultato_lordo_gestione = load_matrix(
        RAW_DIR / "SIMM_TOT_63BxB.xlsx", f"STOTBB_{YEAR}"
    )
    imp_codes, _, imp_matrix, _, _, _, _ = load_matrix(
        RAW_DIR / "SIMM_IMP_63BxB.xlsx", f"SIMPBB_{YEAR}"
    )
    household_consumption, hours_worked_thousands = load_use_table_columns(
        RAW_DIR / "USEPB_63B.xlsx", f"uspb{str(YEAR)[2:]}"
    )

    assert tot_codes == imp_codes, "Le due tavole devono avere le stesse branche nello stesso ordine"
    n = len(tot_codes)

    # Matrice domestica = totale (domestico + import) meno la sola componente import.
    domestic_matrix = [
        [tot_matrix[i][j] - imp_matrix[i][j] for j in range(n)]
        for i in range(n)
    ]

    # Coefficienti tecnici domestici: A[i][j] = flusso domestico i->j / produzione branca j
    tech_coeff = [[0.0] * n for _ in range(n)]
    for j in range(n):
        out_j = output[j]
        if out_j:
            for i in range(n):
                tech_coeff[i][j] = domestic_matrix[i][j] / out_j

    va_coeff = [
        (value_added[j] / output[j]) if output[j] else 0.0
        for j in range(n)
    ]

    # Riga "famiglie" della SAM: reddito da lavoro dipendente generato per unità di
    # produzione di ciascuna branca.
    household_income_coeff = [
        (redditi_lavoro[j] / output[j]) if output[j] else 0.0
        for j in range(n)
    ]

    # Colonna "famiglie" della SAM: come le famiglie allocano un euro aggiuntivo di
    # spesa indotta tra le 63 branche (mix osservato della spesa per consumi finali
    # delle famiglie), moltiplicato per la propensione marginale al consumo (quota che
    # diventa nuova spesa, il resto è cuneo fiscale/contributivo + risparmio).
    hh_total = sum(household_consumption)
    household_spending_coeff = [
        (household_consumption[i] / hh_total) * PROPENSITA_MARGINALE_CONSUMO if hh_total else 0.0
        for i in range(n)
    ]

    # Coefficiente occupazionale: ore lavorate (migliaia) per unità di produzione
    # (milioni di euro) di ciascuna branca.
    hours_coeff = [
        (hours_worked_thousands[j] / output[j]) if output[j] else 0.0
        for j in range(n)
    ]

    cache = {
        "source": "ISTAT - Sistema di tavole Input-Output, anni 2020-2022 "
                   "(https://www.istat.it/tavole-di-dati/il-sistema-di-tavole-input-output-anni-2020-2022/); "
                   "USEPB_63B.xlsx per spesa consumi famiglie e ore lavorate; "
                   "cuneo fiscale OECD Taxing Wages 2023 (Italia, lavoratore medio, 2022: 45,0%); "
                   "propensione al risparmio Istat 'Conti nazionali per settore istituzionale 1995-2023' "
                   "(5/4/2024, pag.4: 7,8% nel 2022).",
        "year": YEAR,
        "classification": "NACE Rev.2 / CPA ver.2.1, 63 branche",
        "branch_codes": tot_codes,
        "branch_names": tot_names,
        "technical_coefficients_domestic": tech_coeff,
        "value_added_coefficients": va_coeff,
        "household_income_coefficients": household_income_coeff,
        "household_spending_coefficients": household_spending_coeff,
        "hours_worked_coefficients_thousands_per_million_eur": hours_coeff,
        "output_eur": output,
        "value_added_eur": value_added,
        "gross_operating_surplus_eur": risultato_lordo_gestione,
        "tax_wedge_labour_2022": TAX_WEDGE_LAVORO_2022,
        "household_savings_rate_2022": PROPENSIONE_RISPARMIO_FAMIGLIE_2022,
        "marginal_propensity_to_spend": PROPENSITA_MARGINALE_CONSUMO,
        "note": "Coefficienti tecnici calcolati sulla sola componente domestica (tavola simmetrica "
                "totale meno tavola simmetrica delle importazioni), per un moltiplicatore 'tipo I' "
                "che non sovrastimi l'impatto includendo input importati. I coefficienti famiglie "
                "(household_income_coefficients/household_spending_coefficients) estendono il modello "
                "a una SAM 'tipo II' con il giro indotto dei consumi (vedi multipliers.py). Ipotesi "
                "semplificativa dichiarata: si usa il cuneo fiscale medio nazionale e la propensione al "
                "risparmio media come proxy delle propensioni MARGINALI (prassi standard nei modelli SAM "
                "semplici, ma resta un'approssimazione).",
    }

    out_path = DATA_DIR / f"istat_io_{YEAR}.json"
    out_path.write_text(json.dumps(cache, ensure_ascii=False, indent=None))
    print(f"Scritto {out_path} ({n} branche)")


if __name__ == "__main__":
    main()
